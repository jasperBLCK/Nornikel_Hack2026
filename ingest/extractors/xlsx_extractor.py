"""XLS/XLSX extractor — one Page per sheet, rows rendered pipe-separated.

Legacy .xls files are converted to .xlsx via LibreOffice (headless) first.
"""
from __future__ import annotations

import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from ingest.extractors.base import Extractor
from ingest.structures import Page

_MAX_ROWS_PER_SHEET = 20000


def _convert_xls_to_xlsx(path: Path, out_dir: Path) -> Path:
    subprocess.run(
        ["soffice", "--headless", "--convert-to", "xlsx",
         "--outdir", str(out_dir), str(path)],
        check=True, capture_output=True, timeout=300,
    )
    converted = out_dir / (path.stem + ".xlsx")
    if not converted.exists():
        raise FileNotFoundError(f"LibreOffice did not produce {converted}")
    return converted


def _sheet_to_text(ws) -> str:
    lines: list[str] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= _MAX_ROWS_PER_SHEET:
            lines.append(f"[... truncated at {_MAX_ROWS_PER_SHEET} rows]")
            break
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            lines.append(" | ".join(cells))
    return "\n".join(lines)


class XlsxExtractor(Extractor):
    def extract(self, path: Path) -> list[Page]:
        try:
            if path.suffix.lower() == ".xls":
                with tempfile.TemporaryDirectory() as td:
                    xlsx = _convert_xls_to_xlsx(path, Path(td))
                    return self._extract_xlsx(xlsx)
            return self._extract_xlsx(path)
        except Exception:
            return [Page(page_number=1, text="")]

    def _extract_xlsx(self, path: Path) -> list[Page]:
        wb = load_workbook(str(path), read_only=True, data_only=True)
        pages: list[Page] = []
        try:
            for idx, ws in enumerate(wb.worksheets, 1):
                text = _sheet_to_text(ws)
                if text:
                    text = f"[SHEET {ws.title}]\n{text}"
                pages.append(Page(page_number=idx, text=text))
        finally:
            wb.close()
        return pages or [Page(page_number=1, text="")]
